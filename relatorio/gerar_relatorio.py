#!/usr/bin/env python3
"""
relatorio/gerar_relatorio.py

Gera relatório XLSX a partir do validacao_*.json.

Colunas: EMPRESA | CNPJ | [evento-tabela ...]

Célula por evento:
  - CORRETO          → estava correto (verde)
  - id_rubrica       → foi corrigido (laranja)
  - N/A              → encontrado no holerite mas não localizado no eSocial (cinza)
  - ERRO             → falhou ao corrigir (vermelho)
  - (vazio)          → evento não se aplica a esta empresa

Uso:
  python relatorio/gerar_relatorio.py [validacao_*.json]
"""

import csv
import glob
import json
import os
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR      = Path(__file__).parent.parent
DADOS_ENTRADA = Path(os.environ.get("DADOS_ENTRADA", BASE_DIR / "dados" / "entrada"))
DADOS_SAIDA   = Path(os.environ.get("DADOS_SAIDA",   BASE_DIR / "dados" / "saida"))

# ---------------------------------------------------------------------------
# Cores
# ---------------------------------------------------------------------------
COR_HEADER     = PatternFill("solid", fgColor="2E4057")   # azul escuro
COR_CORRETO    = PatternFill("solid", fgColor="C6EFCE")   # verde claro
COR_CORRIGIDO  = PatternFill("solid", fgColor="FFEB9C")   # amarelo/laranja
COR_NA         = PatternFill("solid", fgColor="EDEDED")   # cinza claro
COR_ERRO       = PatternFill("solid", fgColor="FFC7CE")   # vermelho claro
COR_DEMISSAO   = PatternFill("solid", fgColor="FCE4D6")   # laranja claro (coluna demissão)
COR_EMPRESA    = PatternFill("solid", fgColor="D9E1F2")   # azul claro

FONTE_HEADER   = Font(bold=True, color="FFFFFF", size=10)
FONTE_EMPRESA  = Font(bold=True, size=10)
FONTE_NORMAL   = Font(size=10)

ALINHAMENTO_C  = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALINHAMENTO_E  = Alignment(horizontal="left",   vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------
# Carregar eventos do esocial.csv
# ---------------------------------------------------------------------------

def carregar_eventos() -> list[dict]:
    """
    Retorna lista de eventos únicos por (id_evento, tabela), preservando ordem.
    Cada item: {id_evento, nome_esocial, tabela, demissao, col_key}
    """
    eventos = []
    vistos = set()
    with open(DADOS_ENTRADA / "esocial.csv", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter=";")
        for row in reader:
            id_ev  = row["id_evento"].strip()
            tabela = row["tabela"].strip()
            nome   = row["nome_esocial"].strip()
            demissao = row.get("demissão", "Não").strip().lower() in ("sim", "s", "1", "true")
            key = (id_ev, tabela)
            if key not in vistos:
                vistos.add(key)
                eventos.append({
                    "id_evento": id_ev,
                    "nome_esocial": nome,
                    "tabela": tabela,
                    "demissao": demissao,
                    "col_key": key,
                })
    return eventos


# ---------------------------------------------------------------------------
# Construir lookup por empresa: (id_evento, tabela) → resultado
# ---------------------------------------------------------------------------

def _nome_para_id(eventos: list[dict]) -> dict:
    """Mapeia nome_esocial → id_evento (primeira ocorrência)."""
    m = {}
    for ev in eventos:
        m.setdefault(ev["nome_esocial"].upper(), ev["id_evento"])
    return m


def _parse_nao_encontrado(texto: str) -> tuple[str, str]:
    """'136 (Holerite)' → ('136', 'Holerite')"""
    m = re.match(r"^(\d+)\s*\((.+)\)$", texto.strip())
    if m:
        return m.group(1), m.group(2)
    return texto.strip(), ""


def construir_lookup(empresa: dict, eventos: list[dict]) -> dict:
    """
    Retorna dict: (id_evento, tabela) → {'status': str, 'id_rubrica': str}
    """
    nome_para_id = _nome_para_id(eventos)
    resultado = {}

    # Rubricas encontradas e validadas
    for r in empresa.get("rubricas", []):
        nome_ev = r.get("nome_evento", "").upper()
        tabela  = r.get("tabela", "")
        id_ev   = nome_para_id.get(nome_ev)
        if not id_ev:
            continue
        key = (id_ev, tabela)
        status = r.get("status", "")
        resultado[key] = {
            "status":     status,
            "id_rubrica": r.get("id_rubrica", ""),
        }

    # Não encontrados no eSocial
    for texto in empresa.get("nao_encontrados", []):
        id_ev, tabela = _parse_nao_encontrado(texto)
        key = (id_ev, tabela)
        if key not in resultado:
            resultado[key] = {"status": "N/A", "id_rubrica": ""}

    return resultado


# ---------------------------------------------------------------------------
# Formatar valor da célula
# ---------------------------------------------------------------------------

def celula(lookup: dict, col_key: tuple) -> tuple[str, PatternFill]:
    """Retorna (texto, fill) para a célula do evento."""
    if col_key not in lookup:
        return "", None

    info   = lookup[col_key]
    status = info["status"]
    id_r   = info["id_rubrica"]

    if status in ("CORRETO", "CORRIGIDO_EXTERNAMENTE"):
        return "CORRETO", COR_CORRETO

    if status == "CORRIGIDO":
        return id_r or "CORRIGIDO", COR_CORRIGIDO

    if status == "N/A":
        return "N/A", COR_NA

    if status in ("ERRADO", "ERRO_FORM", "ERRO_ASSINATURA"):
        return f"ERRO\n{id_r}" if id_r else "ERRO", COR_ERRO

    return status, None


# ---------------------------------------------------------------------------
# Gerar XLSX
# ---------------------------------------------------------------------------

def gerar_xlsx(validacao: dict, eventos: list[dict], output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"

    # --- Cabeçalho ---
    headers_fixos = ["EMPRESA", "CNPJ"]
    ws.append(headers_fixos + [
        f"{ev['id_evento']} - {ev['nome_esocial']}\n({ev['tabela']})"
        for ev in eventos
    ])

    header_row = ws[1]
    for i, cell in enumerate(header_row):
        cell.font      = FONTE_HEADER
        cell.fill      = COR_HEADER if i >= 2 else COR_HEADER
        cell.alignment = ALINHAMENTO_C
        # Colunas de demissão com fundo levemente diferente
        if i >= 2 and eventos[i - 2]["demissao"]:
            cell.fill = PatternFill("solid", fgColor="8B3A3A")

    ws.row_dimensions[1].height = 45

    # --- Dados ---
    for cnpj, empresa in sorted(validacao.items(), key=lambda x: x[1].get("nome", "")):
        nome   = empresa.get("nome", "")
        lookup = construir_lookup(empresa, eventos)

        row_values = [nome, cnpj]
        row_fills  = [COR_EMPRESA, COR_EMPRESA]

        for ev in eventos:
            texto, fill = celula(lookup, ev["col_key"])
            row_values.append(texto)
            row_fills.append(fill)

        ws.append(row_values)

        linha = ws.max_row
        ws.row_dimensions[linha].height = 18
        for col_idx, (valor, fill) in enumerate(zip(row_values, row_fills), start=1):
            c = ws.cell(row=linha, column=col_idx)
            c.font      = FONTE_EMPRESA if col_idx <= 2 else FONTE_NORMAL
            c.alignment = ALINHAMENTO_E if col_idx <= 2 else ALINHAMENTO_C
            if fill:
                c.fill = fill

    # --- Larguras ---
    ws.column_dimensions["A"].width = 40  # EMPRESA
    ws.column_dimensions["B"].width = 18  # CNPJ
    for i in range(len(eventos)):
        col_letter = get_column_letter(i + 3)
        ws.column_dimensions[col_letter].width = 18

    # --- Congelar cabeçalho ---
    ws.freeze_panes = "C2"

    # --- Aba de legenda ---
    ws_leg = wb.create_sheet("Legenda")
    legenda = [
        ("CORRETO",  COR_CORRETO,   "Rubrica estava correta, nenhuma alteração necessária"),
        ("CORRIGIDO (id_rubrica)", COR_CORRIGIDO, "Rubrica estava errada e foi corrigida — exibe o ID da rubrica"),
        ("N/A",      COR_NA,        "Evento encontrado no holerite mas não localizado no eSocial"),
        ("ERRO",     COR_ERRO,      "Falha ao tentar corrigir a rubrica"),
        ("(vazio)",  None,          "Evento não se aplica a esta empresa"),
    ]
    ws_leg.append(["Status", "Cor", "Significado"])
    ws_leg[1][0].font = FONTE_HEADER; ws_leg[1][0].fill = COR_HEADER
    ws_leg[1][1].font = FONTE_HEADER; ws_leg[1][1].fill = COR_HEADER
    ws_leg[1][2].font = FONTE_HEADER; ws_leg[1][2].fill = COR_HEADER
    ws_leg.column_dimensions["A"].width = 25
    ws_leg.column_dimensions["B"].width = 12
    ws_leg.column_dimensions["C"].width = 60

    for status, fill, desc in legenda:
        ws_leg.append([status, "", desc])
        row = ws_leg.max_row
        if fill:
            ws_leg.cell(row=row, column=1).fill = fill
            ws_leg.cell(row=row, column=2).fill = fill
        for col in range(1, 4):
            ws_leg.cell(row=row, column=col).alignment = ALINHAMENTO_E

    wb.save(output_path)
    print(f"[OK] Relatório salvo em: {output_path}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    # Localizar validacao JSON
    if len(sys.argv) > 1:
        validacao_path = sys.argv[1]
    else:
        arquivos = sorted(glob.glob(str(DADOS_SAIDA / "validacao_*.json")))
        if not arquivos:
            print("[ERRO] Nenhum validacao_*.json encontrado em dados/saida/")
            sys.exit(1)
        validacao_path = arquivos[-1]

    print(f"[Validacao] {validacao_path}")

    with open(validacao_path, encoding="utf-8") as f:
        validacao = json.load(f)

    eventos = carregar_eventos()
    print(f"[Eventos]   {len(eventos)} colunas ({len([e for e in eventos if e['demissao']])} de demissão)")
    print(f"[Empresas]  {len(validacao)} no relatório")

    # Nome do arquivo de saída
    nome_base = Path(validacao_path).stem.replace("validacao", "relatorio")
    output_path = str(DADOS_SAIDA / f"{nome_base}.xlsx")

    gerar_xlsx(validacao, eventos, output_path)


if __name__ == "__main__":
    main()
