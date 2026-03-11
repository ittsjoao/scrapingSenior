"""
Converte o JSON de auditoria (dados/saida/auditoria_*.json) em planilha Excel.

Uso:
    python json_para_xlsx.py                          # pega o JSON mais recente
    python json_para_xlsx.py auditoria_20260310.json  # arquivo específico
"""
import json
import os
import sys
from glob import glob

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

_DIR_SAIDA = os.path.join(os.path.dirname(__file__), "..", "dados", "saida")


def _descobrir_json(caminho_arg=None):
    if caminho_arg:
        if os.path.isabs(caminho_arg):
            return caminho_arg
        return os.path.join(_DIR_SAIDA, caminho_arg)
    arquivos = sorted(glob(os.path.join(_DIR_SAIDA, "auditoria_*.json")))
    if not arquivos:
        sys.exit("Nenhum auditoria_*.json encontrado em " + _DIR_SAIDA)
    return arquivos[-1]


_FILLS = {
    "CORRETO": PatternFill("solid", fgColor="C6EFCE"),
    "ERRADO":  PatternFill("solid", fgColor="FFC7CE"),
    "N/A":     PatternFill("solid", fgColor="D9D9D9"),
}
_FONTS = {
    "ERRADO": Font(color="9C0006", bold=True),
}


def gerar_xlsx(caminho_json):
    with open(caminho_json, encoding="utf-8") as f:
        dados = json.load(f)

    # Coleta todos os nomes de evento na ordem de primeira aparição
    eventos_ordenados = []
    visto = set()
    for info in dados.values():
        for r in info["rubricas"]:
            nome = r["nome_evento"]
            if nome not in visto:
                eventos_ordenados.append(nome)
                visto.add(nome)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Auditoria"

    # --- cabeçalho ---
    cabecalho = ["EMPRESA", "CNPJ"] + eventos_ordenados
    ws.append(cabecalho)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # --- dados ---
    for cnpj, info in dados.items():
        status_map = {}
        for r in info["rubricas"]:
            nome = r["nome_evento"]
            st = r["status"]
            motivo = r.get("motivo") or ""
            valor = f"N/A ({motivo})" if st == "N/A" and motivo else st
            # Se mesmo evento aparece mais de uma vez, prioriza ERRADO > CORRETO > N/A
            anterior = status_map.get(nome)
            if anterior is None or _prioridade(st) > _prioridade(_status_base(anterior)):
                status_map[nome] = valor

        linha = [info["nome"], cnpj]
        for ev in eventos_ordenados:
            linha.append(status_map.get(ev, ""))
        ws.append(linha)

    # --- formatar células de status ---
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=len(cabecalho)):
        for cell in row:
            base = _status_base(cell.value or "")
            if base in _FILLS:
                cell.fill = _FILLS[base]
            if base in _FONTS:
                cell.font = _FONTS[base]
            cell.alignment = Alignment(horizontal="center")

    # --- largura das colunas ---
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 22
    for col_letter in [openpyxl.utils.get_column_letter(i) for i in range(3, len(cabecalho) + 1)]:
        ws.column_dimensions[col_letter].width = 18

    # --- salvar ---
    nome_base = os.path.splitext(os.path.basename(caminho_json))[0]
    caminho_xlsx = os.path.join(_DIR_SAIDA, f"{nome_base}.xlsx")
    wb.save(caminho_xlsx)
    print(f"Planilha salva: {caminho_xlsx}")
    print(f"  {len(dados)} empresas × {len(eventos_ordenados)} eventos")
    return caminho_xlsx


def _prioridade(status):
    return {"ERRADO": 2, "CORRETO": 1, "N/A": 0}.get(status, -1)


def _status_base(valor):
    if not valor:
        return ""
    if valor.startswith("N/A"):
        return "N/A"
    return valor


if __name__ == "__main__":
    arg = sys.argv[1] if len(sys.argv) > 1 else None
    caminho = _descobrir_json(arg)
    print(f"Lendo: {caminho}")
    gerar_xlsx(caminho)
