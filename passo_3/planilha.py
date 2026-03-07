# passo_3/planilha.py
import openpyxl


def carregar_dados(caminho):
    """
    Lê a planilha de entrada e retorna estrutura agrupada por CNPJ.

    Retorna dict:
    {
        cnpj: {
            "eventos": {
                (evento, evento_aux, irrf, tabela): [
                    {"cpf": str, "competencia": str},
                    ...
                ]
            },
            "todos_cpfs": [str, ...]  # CPFs únicos, ordem de aparição
        }
    }

    Linhas com DEMISSÃO=SIM → status "Demissão" (puladas).
    Linhas com CPF vazio → status "CPF em branco" (puladas).
    Ambas são retornadas em lista separada para registro no resultado.
    """
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb.active

    headers = [str(c.value).strip().upper() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

    def col(row, nome):
        idx = headers.index(nome)
        val = row[idx].value
        return str(val).strip() if val is not None else ""

    empresas = {}   # cnpj → {eventos: {}, todos_cpfs: []}
    puladas = []    # lista de dicts com status já definido

    for row in ws.iter_rows(min_row=2):
        cnpj       = col(row, "CNPJ")
        evento     = col(row, "EVENTO")
        evento_aux = col(row, "EVENTO_AUX")
        irrf       = col(row, "IRRF")
        tabela     = col(row, "TABELA")
        demissao   = col(row, "DEMISSÃO").upper()
        cpf        = col(row, "CPF")
        competencia = col(row, "COMPETENCIA")

        if not cnpj:
            continue

        if demissao == "SIM":
            puladas.append({
                "cnpj": cnpj, "evento": evento, "irrf": irrf,
                "cpf": cpf, "competencia": competencia, "status": "Demissão"
            })
            continue

        if not cpf:
            puladas.append({
                "cnpj": cnpj, "evento": evento, "irrf": irrf,
                "cpf": cpf, "competencia": competencia, "status": "CPF em branco"
            })
            continue

        if cnpj not in empresas:
            empresas[cnpj] = {"eventos": {}, "todos_cpfs": []}

        chave = (evento, evento_aux, irrf, tabela)
        if chave not in empresas[cnpj]["eventos"]:
            empresas[cnpj]["eventos"][chave] = []
        empresas[cnpj]["eventos"][chave].append({"cpf": cpf, "competencia": competencia})

        if cpf not in empresas[cnpj]["todos_cpfs"]:
            empresas[cnpj]["todos_cpfs"].append(cpf)

    return empresas, puladas
