# passo_2/criar_planilha.py
import os
import csv
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ENTRADA_DIR = os.path.join(BASE_DIR, "dados", "entrada")
SAIDA_DIR   = os.path.join(BASE_DIR, "dados", "saida")
OUTPUT_PATH = os.path.join(BASE_DIR, "passo_2", "planilha_empresas.xlsx")


def ler_esocial():
    """Retorna lista de dicts na ordem do CSV."""
    path = os.path.join(ENTRADA_DIR, "esocial.csv")
    rows = []
    with open(path, encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter=";")
        for row in reader:
            if not row.get("id_evento", "").strip():
                continue
            rows.append({
                "id_evento":    int(row["id_evento"].strip()),
                "nome_esocial": row["nome_esocial"].strip(),
                "irf":          row["irf"].strip(),
                "tabela":       row["tabela"].strip(),
                "demissao":     row["demissão"].strip(),
            })
    return rows

def ler_eventos():
    """Retorna dict {id_evento (int): nome_evento (str)}."""
    path = os.path.join(ENTRADA_DIR, "eventos.csv")
    eventos = {}
    with open(path, encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter=";")
        for row in reader:
            if not row.get("id_evento", "").strip():
                continue
            eventos[int(row["id_evento"].strip())] = row["nome_evento"].strip()
    return eventos

def ler_empresas():
    """Retorna dict {nome_empresa (str): id_empresa (str)}."""
    path = os.path.join(ENTRADA_DIR, "empresas.csv")
    empresas = {}
    with open(path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter=";")
        for row in reader:
            nome = row.get("nome_empresa", "").strip()
            id_  = row.get("id_empresa", "").strip()
            if nome and id_:
                empresas[nome] = id_
    return empresas

COLLAB_RE = re.compile(
    r'^\s{5,}'
    r'([A-ZÁÉÍÓÚÃÕÂÊÔÀÇÜ][A-ZÁÉÍÓÚÃÕÂÊÔÀÇÜa-záéíóúãõâêôàçü\s\.\-]+?)'
    r'\s{2,}(\d{3})\s+'
    r'(\d{2}/\d{4})'
)

IGNORE_PATTERNS = [
    re.compile(r'^\s*\d+\s+-\s+'),        # cabeçalho empresa
    re.compile(r'^\s+\d{4}\s+-'),         # código evento
    re.compile(r'Total de Colaboradores'),
    re.compile(r'FPRF004'),
    re.compile(r'Per[íi]odo:'),
    re.compile(r'Tipo:'),
    re.compile(r'Evento\s+Colaborador'),
    re.compile(r'^\s*$'),                  # linha vazia
    re.compile(r'P[áa]g\.'),              # página
]


def _ler_txt(path):
    """Lê arquivo TXT tentando CP1252 depois latin-1."""
    for enc in ("cp1252", "latin-1"):
        try:
            with open(path, encoding=enc) as f:
                return f.readlines()
        except UnicodeDecodeError:
            continue
    return []


def parsear_txt(path):
    """
    Retorna lista de dicts {colaborador, competencia}, deduplicados por colaborador.
    Mantém apenas a primeira ocorrência de cada colaborador.
    """
    if not os.path.exists(path):
        return []

    linhas = _ler_txt(path)
    vistos = {}  # colaborador → competencia (primeira ocorrência)

    for linha in linhas:
        if any(p.search(linha) for p in IGNORE_PATTERNS):
            continue
        m = COLLAB_RE.match(linha)
        if m:
            nome = " ".join(m.group(1).split())  # normaliza espaços internos
            competencia = m.group(3)
            if nome not in vistos:
                vistos[nome] = competencia

    return [{"colaborador": k, "competencia": v} for k, v in vistos.items()]

def gerar_excel(esocial_rows, eventos, empresas, pastas_existentes):
    wb = Workbook()
    ws = wb.active
    ws.title = "Planilha"

    # Cabeçalho
    cabecalho = ["ID SENIOR", "EMPRESA", "EVENTO", "IRRF", "COLABORADOR", "COMPETENCIA"]
    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    for col, titulo in enumerate(cabecalho, 1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.font = header_font
        cell.fill = header_fill

    pastas_set = set(pastas_existentes)
    # Universo de empresas: union pastas + empresas.csv
    todas = sorted(pastas_set | set(empresas.keys()))
    com_pasta = [e for e in todas if e in pastas_set]
    sem_pasta = [e for e in todas if e not in pastas_set]

    fonte_vermelha = Font(color="FF0000")
    linha_atual = 2

    for empresa in com_pasta + sem_pasta:
        tem_pasta = empresa in pastas_set
        id_senior = empresas.get(empresa, "")

        for esocial_row in esocial_rows:
            id_evento    = esocial_row["id_evento"]
            nome_esocial = esocial_row["nome_esocial"]
            irf          = esocial_row["irf"]
            nome_evento  = eventos.get(id_evento)

            if nome_evento and tem_pasta:
                txt_path = os.path.join(SAIDA_DIR, empresa, nome_evento + ".TXT")
                colaboradores = parsear_txt(txt_path)
            else:
                colaboradores = []

            linhas_evento = colaboradores if colaboradores else [{"colaborador": "", "competencia": ""}]

            for colab in linhas_evento:
                valores = [
                    id_senior,
                    empresa,
                    nome_esocial,
                    irf,
                    colab["colaborador"],
                    colab["competencia"],
                ]
                for col, val in enumerate(valores, 1):
                    cell = ws.cell(row=linha_atual, column=col, value=val)
                    if not tem_pasta:
                        cell.font = fonte_vermelha
                linha_atual += 1

    # Ajuste de largura de colunas
    larguras = [12, 45, 45, 8, 50, 14]
    for col, largura in enumerate(larguras, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = largura

    wb.save(OUTPUT_PATH)

if __name__ == "__main__":
    esocial_rows       = ler_esocial()
    eventos            = ler_eventos()
    empresas           = ler_empresas()
    pastas_existentes  = sorted(os.listdir(SAIDA_DIR))
    gerar_excel(esocial_rows, eventos, empresas, pastas_existentes)
    print("Planilha gerada:", OUTPUT_PATH)
